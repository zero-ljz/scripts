import json
from datetime import datetime, date
from openpyxl import load_workbook


class ExcelToListSerializer:
    def __init__(self, file_path):
        self.file_path = file_path
        # data_only=True 确保读取的是值而不是公式
        self.wb = load_workbook(file_path, data_only=True)

    def _convert_value(self, value):
        """处理 JSON 不支持的 Excel 对象类型"""
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        # 如果是布尔值、数字或 None，保持原样；其他转为字符串
        if value is None or isinstance(value, (int, float, bool)):
            return value
        return str(value)

    def serialize(self):
        result = {}

        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            # 使用 values 迭代器，直接按行获取数据，不跳过首行
            sheet_matrix = []

            for row in sheet.values:
                # 处理当前行中的每一个单元格
                processed_row = [self._convert_value(cell) for cell in row]
                sheet_matrix.append(processed_row)

            result[sheet_name] = sheet_matrix

        return result

    def save_to_json(self, output_path):
        data = self.serialize()
        with open(output_path, "w", encoding="utf-8") as f:
            # ensure_ascii=False 保证中文字符正常显示
            json.dump(data, f, ensure_ascii=False, indent=4)
        print(f"✅ 转换完成，已保存至: {output_path}")


# --- 快速运行 ---
if __name__ == "__main__":
    # 填入你的文件路径
    input_file = "1.xlsx"
    output_file = "1.json"

    serializer = ExcelToListSerializer(input_file)
    serializer.save_to_json(output_file)
