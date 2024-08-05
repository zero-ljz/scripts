import sys, os, datetime
from openpyxl import load_workbook

filename = ""
wb = load_workbook(filename = filename)

if 'Sheet1' in wb.sheetnames:
    ws = wb['Sheet1']
else:
    ws = wb.active
# print(ws['D18'].value)

data = []
first_row_values = tuple(cell.value for cell in ws[1])
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False)): # 从第二行开始读取
    row_num = row[0].row
    row_values = tuple(cell.value for cell in row)
    # print('i:', i, 'row_num:', row_num, 'row_values:', row_values)
    if row_values[0] is None: # 第一列为空时跳过
        continue

    print('当前导入Excel第', row_num, '行')
    row_dict = {k: v for k, v in zip(first_row_values, row_values)} # 将首行（表头行）和当前行合并为字典
    data.append(row_dict)