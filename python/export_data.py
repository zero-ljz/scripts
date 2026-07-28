

import pymysql
import json
import datetime
import argparse
import os
from openpyxl import Workbook, load_workbook

# --- 数据库配置 (请根据实际情况修改) ---
DB_CONFIG = {
    "host": "localhost",
    "port": 3306,
    "user": "root",
    "password": "123123",
    "database": "",
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor  # 返回字典格式，方便处理
}

exclude_tables = []

# 处理 JSON 序列化中的日期时间问题
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime.datetime, datetime.date)):
            return obj.isoformat()
        return super(DateTimeEncoder, self).default(obj)

def get_conn():
    return pymysql.connect(**DB_CONFIG)

def export_data(filename):
    conn = get_conn()
    try:
        with conn.cursor() as cursor:
            # 获取所有表名
            cursor.execute("SHOW TABLES")
            all_tables = [list(row.values())[0] for row in cursor.fetchall()]
            tables = [t for t in all_tables if t not in exclude_tables]
            
            all_data = {}
            # 创建 Excel 工作簿
            wb = Workbook()
            wb.remove(wb.active) # 删除默认 sheet

            sql_lines = []

            for table in tables:
                cursor.execute(f"SELECT * FROM `{table}`")
                rows = cursor.fetchall() # 获取的是字典列表
                all_data[table] = rows

                # 1. 准备 Excel Sheet
                ws = wb.create_sheet(title=table)
                if rows:
                    headers = list(rows[0].keys())
                    ws.append(headers)
                    for row in rows:
                        ws.append(list(row.values()))

                # 2. 准备 SQL 脚本内容
                for row in rows:
                    cols = ", ".join([f"`{k}`" for k in row.keys()])
                    # 处理 SQL 中的 None 为 NULL，并对字符串转义
                    vals = []
                    for v in row.values():
                        if v is None: vals.append("NULL")
                        elif isinstance(v, (int, float)): vals.append(str(v))
                        else: vals.append(f"'{conn.escape_string(str(v))}'")
                    sql_lines.append(f"REPLACE INTO `{table}` ({cols}) VALUES ({', '.join(vals)});")

            # 保存 JSON
            with open(f"{filename}.json", "w", encoding="utf-8") as f:
                json.dump(all_data, f, cls=DateTimeEncoder, ensure_ascii=False, indent=4, default=str)
            
            # 保存 Excel
            wb.save(f"{filename}.xlsx")

            # 保存 SQL
            with open(f"{filename}.sql", "w", encoding="utf-8") as f:
                f.write("\n".join(sql_lines))
            
            print(f"√ 导出成功: {filename}.json/xlsx/sql")
    finally:
        conn.close()

def import_data(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    conn = get_conn()
    try:
        with conn.cursor() as cursor:
            # 禁用外键检查 ---
            cursor.execute("SET FOREIGN_KEY_CHECKS = 0;")
            
            if ext == ".json":
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for table, rows in data.items():
                    for row in rows:
                        cols = ", ".join([f"`{k}`" for k in row.keys()])
                        placeholders = ", ".join(["%s"] * len(row))
                        # 使用 REPLACE INTO 避免主键冲突，或者先清空表
                        cursor.execute(f"REPLACE INTO `{table}` ({cols}) VALUES ({placeholders})", list(row.values()))
            
            elif ext == ".xlsx":
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.values)
                    if not rows: continue
                    headers = rows[0]
                    for row in rows[1:]:
                        cols = ", ".join([f"`{k}`" for k in headers])
                        placeholders = ", ".join(["%s"] * len(row))
                        cursor.execute(f"REPLACE INTO `{sheet_name}` ({cols}) VALUES ({placeholders})", row)

            elif ext == ".sql":
                with open(file_path, "r", encoding="utf-8") as f:
                    statements = f.read().split(";")
                    for stmt in statements:
                        if stmt.strip(): cursor.execute(stmt)
                        
            # 恢复外键检查 ---
            cursor.execute("SET FOREIGN_KEY_CHECKS = 1;")
            conn.commit()
            print(f"√ 导入 {file_path} 成功")
    finally:
        conn.close()

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("action", choices=["export", "import"])
    parser.add_argument("-f", "--file", required=True)
    args = parser.parse_args()

    if args.action == "export":
        export_data(args.file)
    else:
        import_data(args.file)